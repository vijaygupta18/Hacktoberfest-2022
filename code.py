#BILL GENERATOR IN MS-EXCEL USING OPENPYXL LIBRARY OF PYTHON
#Coded by Deepanshu Mittal
#python3

import openpyxl             #install openpyxl using code: pip install openpyxl
from datetime import datetime       #import datetime module for time at the time of billing
from openpyxl.styles import Alignment       #just to give a proper alignment to the data in MS Excel

wb = openpyxl.Workbook()     #create a new excel file 
sheet = wb.active

today=datetime.now()        #time during billing

sheet.merge_cells('A1:E1')
sheet['A1']="BILL GENERATOR"
sheet['A1'].alignment = Alignment(horizontal='center')

sheet.merge_cells('A2:B2')
sheet.merge_cells('D2:E2')
sheet['A2']="Deepanshu Mittal"          #Customer name can be be here
sheet['A2'].alignment = Alignment(horizontal='left')
sheet['D2']=today
sheet['D2'].alignment = Alignment(horizontal='right')

#Give headings of the Bill Table
sheet['A4']="#"
sheet['B4']="Item"
sheet['C4']="Quantity"
sheet['D4']="Price per 1 unit"
sheet['E4']="Amount"

#----------IMPORTANT---------------------------------
#Format for file containing list of items should be:

# Quantity | Item Name | Price per 1 unit

#------------------------------------------------------

fh=open("list.txt","r+")                #open file to read the list of items
data=fh.readlines()

done={}                 #to keep track of marked items in the bill
i=1
p=i
total_amount=0          #to sum up the total bill
for line in data:
    line = line.strip()
    item = line.split('|')
    qty = float(item[0].strip())
    name = item[1].strip()
    price = float(item[2].strip())
    if(done.get(name,0) == 0):
        #A new item is added to the bill
        sheet.cell(row=i+4, column=1).value = i
        sheet.cell(row=i+4, column=2).value = name
        sheet.cell(row=i+4, column=3).value = qty
        sheet.cell(row=i+4, column=4).value = price
        sheet.cell(row=i+4, column=5).value = qty*price
        done[name] = i+4
        total_amount += qty*price
        i += 1
        p=i
    else:
        #Previous item appears again in the list so it is updated in the bill
        sheet.cell(row=done[name], column=3).value += qty
        sheet.cell(row=done[name], column=5).value += qty*price
        total_amount += qty*price

#Finally the total amount is revealed
sheet.cell(row=p+4, column=4).value = "TOTAL"
sheet.cell(row=p+4, column=5).value = total_amount

#save at last the workbook
wb.save(filename='Bill.xlsx')